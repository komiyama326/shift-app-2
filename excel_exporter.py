# excel_exporter.py

import calendar
import datetime
from dateutil.relativedelta import relativedelta
import holidays
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- 定数定義 ---
FONT_NAME = 'メイリオ'
weekdays_jp = ("月", "火", "水", "木", "金", "土", "日")

# --- メイン関数 ---
def export_to_excel(filepath, year, month, title, schedule_data, staff_manager, prev_month_schedule=None, format_type='grid'):
    """
    シフトデータを指定されたフォーマットでExcelファイルに出力する。
    :param format_type: 'grid' または 'list'
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月シフト"

    if format_type == 'grid':
        _generate_grid_format(ws, year, month, title, schedule_data, staff_manager, prev_month_schedule)
    elif format_type == 'list':
        _generate_list_format(ws, year, month, title, schedule_data, staff_manager)
    else:
        raise ValueError("Unsupported format_type. Must be 'grid' or 'list'.")

    try:
        wb.save(filepath)
        print(f"Excelファイルが正常に出力されました: {filepath}")
        return True, None
    except Exception as e:
        print(f"Excelファイルの出力中にエラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        return False, str(e)


# --- グリッド形式Excelの生成 ---
def _generate_grid_format(ws, year, month, title, schedule_data, staff_manager, prev_month_schedule):
    """従来のグリッド形式のカレンダーを生成する"""
    # スタイル定義
    font_title = Font(name=FONT_NAME, size=18, bold=True)
    font_month = Font(name=FONT_NAME, size=18, bold=True)

    # 曜日ヘッダーのフォント (サイズを18ptに統一)
    font_weekday_white = Font(name=FONT_NAME, size=18, bold=True, color='FFFFFF')
    font_weekday_black = Font(name=FONT_NAME, size=18, bold=True, color='000000')
    
    # 日付のフォント (サイズを18ptに統一)
    font_date = Font(name=FONT_NAME, size=18, bold=True)

    # 担当者名のフォント (前回から変更なし: 20pt)
    font_staff = Font(name=FONT_NAME, size=20, bold=True)
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side = Side(style='thin', color='000000')
    medium_side = Side(style='medium', color='000000')
    
    # セル塗りつぶし色 (土曜の色を変更、日曜は前回指定のまま)
    fill_sat_header = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid') # ドジャーブルー
    fill_sun_header = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') # 純粋な赤
    fill_weekday_header = PatternFill(start_color='EAF1DD', end_color='EAF1DD', fill_type='solid')
    fill_multi_staff = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')

    # 表の開始位置とサイズ
    start_row, start_col = 2, 2
    num_cols = 7
    end_col = start_col + num_cols - 1
    cal = calendar.monthcalendar(year, month)
    
    # ヘッダー
    ws.merge_cells(start_row=start_row, start_column=start_col + 1, end_row=start_row, end_column=end_col)
    ws.cell(row=start_row, column=start_col, value=f"{month}月").font = font_month
    ws.cell(row=start_row, column=start_col).alignment = align_center
    ws.cell(row=start_row, column=start_col + 1, value=title).font = font_title
    ws.cell(row=start_row, column=start_col + 1).alignment = align_center
    ws.row_dimensions[start_row].height = 30
    
    # 曜日ヘッダー
    weekday_row = start_row + 1
    ws.row_dimensions[weekday_row].height = 22
    for i, day_name in enumerate(weekdays_jp):
        cell = ws.cell(row=weekday_row, column=start_col + i)
        ws.column_dimensions[get_column_letter(start_col + i)].width = 16
        cell.value = day_name
        cell.alignment = align_center
        # ★★★★★ 変更点 2: 曜日によってフォントの色を切り替える ★★★★★
        if i == 5: # 土曜日
            cell.fill = fill_sat_header
            cell.font = font_weekday_white
        elif i == 6: # 日曜日
            cell.fill = fill_sun_header
            cell.font = font_weekday_white
        else: # 平日
            cell.fill = fill_weekday_header
            cell.font = font_weekday_black
    
    # カレンダー本体
    jp_holidays = holidays.JP(years=year)
    prev_month_date = datetime.date(year, month, 1) - relativedelta(months=1)
    _, days_in_prev_month = calendar.monthrange(prev_month_date.year, prev_month_date.month)
    
    current_calendar_row = start_row + 2
    for week in cal:
        max_staff_in_week = 0
        for day in week:
            if day != 0:
                num_staff = len(schedule_data.get(datetime.date(year, month, day), []))
                if num_staff > max_staff_in_week:
                    max_staff_in_week = num_staff
        staff_row_height = 30 + (max_staff_in_week - 1) * 25 if max_staff_in_week > 1 else 30

        ws.row_dimensions[current_calendar_row].height = 25
        ws.row_dimensions[current_calendar_row + 1].height = staff_row_height
        ws.row_dimensions[current_calendar_row + 2].height = 30
        
        for col_idx, day in enumerate(week):
            date_cell = ws.cell(row=current_calendar_row, column=start_col + col_idx)
            staff_cell = ws.cell(row=current_calendar_row + 1, column=start_col + col_idx)
            memo_cell = ws.cell(row=current_calendar_row + 2, column=start_col + col_idx)
            
            date_cell.alignment = align_center
            staff_cell.alignment = align_center
            staff_cell.font = font_staff

            is_current_month = (day != 0)
            target_date, staff_list = None, []
            
            if is_current_month:
                target_date = datetime.date(year, month, day)
                staff_list = schedule_data.get(target_date, [])
                date_cell.value = day
            elif prev_month_schedule:
                first_weekday_of_month = datetime.date(year, month, 1).weekday()
                if current_calendar_row == start_row + 2 and col_idx < first_weekday_of_month:
                    day_num = days_in_prev_month - (first_weekday_of_month - col_idx - 1)
                    target_date = datetime.date(prev_month_date.year, prev_month_date.month, day_num)
                    staff_list = prev_month_schedule.get(target_date, [])
                    date_cell.value = day_num

            if staff_list:
                staff_cell.value = "\n".join([s.name for s in staff_list])
                if is_current_month:
                    if len(staff_list) == 1:
                        color_hex = staff_list[0].color_code.lstrip('#')
                        staff_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    else:
                        staff_cell.fill = fill_multi_staff

            date_cell.font = font_date
            if not is_current_month:
                date_cell.font = Font(name=FONT_NAME, size=font_date.size, bold=font_date.bold, color='A9A9A9')
                staff_cell.font = Font(name=font_staff.name, size=font_staff.size, bold=font_staff.bold, color='A9A9A9')
            elif target_date:
                is_holiday_flag = target_date in jp_holidays
                is_saturday = target_date.weekday() == 5
                is_sunday = target_date.weekday() == 6
                if is_sunday or is_holiday_flag:
                    date_cell.fill = fill_sun_header
                    # 仕様: 日付フォントは曜日・祝日を問わず18ptに統一
                    date_cell.font = Font(name=FONT_NAME, size=18, bold=True, color='FFFFFF')
                elif is_saturday:
                    date_cell.fill = fill_sat_header
                    # 仕様: 日付フォントは曜日・祝日を問わず18ptに統一
                    date_cell.font = Font(name=FONT_NAME, size=18, bold=True, color='FFFFFF')

        current_calendar_row += 3
    
    end_row = current_calendar_row - 1
        
    border_full = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border_full

    for row_idx in range(start_row, end_row + 1):
        ws.cell(row_idx, start_col).border = Border(left=medium_side, top=ws.cell(row_idx, start_col).border.top, right=ws.cell(row_idx, start_col).border.right, bottom=ws.cell(row_idx, start_col).border.bottom)
        ws.cell(row_idx, end_col).border = Border(right=medium_side, top=ws.cell(row_idx, end_col).border.top, left=ws.cell(row_idx, end_col).border.left, bottom=ws.cell(row_idx, end_col).border.bottom)
    for col_idx in range(start_col, end_col + 1):
        ws.cell(start_row, col_idx).border = Border(top=medium_side, left=ws.cell(start_row, col_idx).border.left, right=ws.cell(start_row, col_idx).border.right, bottom=ws.cell(start_row, col_idx).border.bottom)
        ws.cell(end_row, col_idx).border = Border(bottom=medium_side, left=ws.cell(end_row, col_idx).border.left, right=ws.cell(end_row, col_idx).border.right, top=ws.cell(end_row, col_idx).border.top)
            
    for col in range(start_col, end_col + 1):
        cell = ws.cell(weekday_row, col)
        cell.border = Border(top=medium_side, bottom=cell.border.bottom, left=cell.border.left, right=cell.border.right)

    for r in range(start_row + 2, end_row + 1, 3):
        if r + 2 <= end_row:
            for col in range(start_col, end_col + 1):
                cell = ws.cell(r + 2, col)
                cell.border = Border(bottom=medium_side, left=cell.border.left, right=cell.border.right, top=cell.border.top)


# --- タイムラインリスト形式Excelの生成 ---
def _generate_list_format(ws, year, month, title, schedule_data, staff_manager):
    # ... (この関数は変更なし) ...
    all_staff = sorted(staff_manager.get_all_staff(), key=lambda s: s.name)
    
    font_main_title = Font(name=FONT_NAME, size=18, bold=True)
    font_month_title = Font(name=FONT_NAME, size=14)
    font_legend_title = Font(name=FONT_NAME, size=11, bold=True)
    font_legend_staff = Font(name=FONT_NAME, size=10)
    font_header = Font(name=FONT_NAME, size=9, bold=True, color='808080')
    font_date_bold = Font(name=FONT_NAME, size=12, bold=True)
    font_weekday = Font(name=FONT_NAME, size=10)
    font_staff_name = Font(name=FONT_NAME, size=11)
    font_saturday = Font(name=FONT_NAME, color='0070C0')
    font_sunday_holiday = Font(name=FONT_NAME, color='C00000')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left_vcenter = Alignment(horizontal='left', vertical='center')
    align_right_vcenter = Alignment(horizontal='right', vertical='center')
    thin_bottom_border = Border(bottom=Side(style='thin', color='BFBFBF'))
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    
    current_row = 1
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    title_cell = ws.cell(current_row, 1, value=title)
    title_cell.font = font_main_title
    title_cell.alignment = align_center
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    month_title_cell = ws.cell(current_row, 1, value=f"{year}年 {month}月")
    month_title_cell.font = font_month_title
    month_title_cell.alignment = align_center
    ws.row_dimensions[current_row].height = 24
    current_row += 2

    ws.cell(current_row, 1, "凡例").font = font_legend_title
    current_row += 1
    
    legend_start_row = current_row
    max_legend_cols = 5
    for i, staff in enumerate(all_staff):
        col_offset = (i % max_legend_cols) * 2
        row_offset = i // max_legend_cols
        
        indicator_cell = ws.cell(legend_start_row + row_offset, 1 + col_offset)
        indicator_cell.value = "■"
        indicator_cell.font = Font(name=FONT_NAME, color=staff.color_code.lstrip('#'))
        indicator_cell.alignment = align_right_vcenter
        ws.column_dimensions[get_column_letter(1 + col_offset)].width = 4

        name_cell = ws.cell(legend_start_row + row_offset, 2 + col_offset)
        name_cell.value = staff.name
        name_cell.font = font_legend_staff
        name_cell.alignment = align_left_vcenter
        ws.column_dimensions[get_column_letter(2 + col_offset)].width = 12
        
    current_row += (len(all_staff) // max_legend_cols) + 2

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(current_row, 1, "DATE & DAY").font = font_header
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=10)
    ws.cell(current_row, 3, "STAFF").font = font_header
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    jp_holidays = holidays.JP(years=year)
    _, num_days = calendar.monthrange(year, month)

    for day_num in range(1, num_days + 1):
        date = datetime.date(year, month, day_num)
        staff_list = schedule_data.get(date, [])
        weekday_idx = date.weekday()
        
        date_cell = ws.cell(current_row, 1, f"{day_num:02d}")
        weekday_cell = ws.cell(current_row, 2, f"({weekdays_jp[weekday_idx]})")
        
        is_saturday = weekday_idx == 5
        is_sunday = weekday_idx == 6
        is_holiday = date in jp_holidays
        
        font_color = "000000"
        if is_sunday or is_holiday: font_color = font_sunday_holiday.color.rgb
        elif is_saturday: font_color = font_saturday.color.rgb

        date_cell.font = Font(name=FONT_NAME, size=12, bold=True, color=font_color)
        weekday_cell.font = Font(name=FONT_NAME, size=10, color=font_color)
        date_cell.alignment = align_right_vcenter
        weekday_cell.alignment = align_left_vcenter
        
        if staff_list:
            staff_per_line = 4
            for i, staff in enumerate(staff_list):
                col_offset = (i % staff_per_line) * 2
                row_offset = i // staff_per_line
                
                indicator_cell = ws.cell(current_row + row_offset, 3 + col_offset)
                indicator_cell.value = "■"
                indicator_cell.font = Font(name=FONT_NAME, size=11, color=staff.color_code.lstrip('#'))
                indicator_cell.alignment = align_right_vcenter
                
                name_cell = ws.cell(current_row + row_offset, 4 + col_offset)
                name_cell.value = staff.name
                name_cell.font = font_staff_name
                name_cell.alignment = align_left_vcenter
        
        num_staff_rows = (len(staff_list) - 1) // staff_per_line + 1 if staff_list else 1
        
        for i in range(num_staff_rows):
            ws.row_dimensions[current_row + i].height = 22
        
        if num_staff_rows > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + num_staff_rows - 1, end_column=1)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + num_staff_rows - 1, end_column=2)
            date_cell.alignment = Alignment(horizontal='right', vertical='top')
            weekday_cell.alignment = Alignment(horizontal='left', vertical='top')

        for col in range(1, 11):
            ws.cell(current_row + num_staff_rows - 1, col).border = thin_bottom_border
        
        current_row += num_staff_rows
